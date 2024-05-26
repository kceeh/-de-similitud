from tkinter import filedialog, Tk, Text, Scrollbar, Menu, messagebox, Button
from openpyxl import load_workbook
import spacy

nlp = spacy.load("es_core_news_md")

def seleccionar_excel():
    root.withdraw()
    messagebox.showinfo("Selección de Archivos", "Por favor, selecciona el archivo Excel con los issues.")

    ruta_excel_1 = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx;*.xlsm"), ("Todos los archivos", "*.*")])

    if ruta_excel_1:
        messagebox.showinfo("Selección de Archivos", "Por favor, selecciona el archivo Excel con los defectos.")

        ruta_excel_2 = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx;*.xlsm"), ("Todos los archivos", "*.*")])

        if ruta_excel_2:
            mostrar_defectos(ruta_excel_1, ruta_excel_2)

def mostrar_defectos(ruta_excel_1, ruta_excel_2):
    workbook_1 = load_workbook(ruta_excel_1)
    workbook_2 = load_workbook(ruta_excel_2)

    hoja_1 = workbook_1['Hoja1']
    hoja_2 = workbook_2['Hoja1']

    descripcion_columna_1 = 8
    descripcion_columna_2 = 2
    numero_defecto_columna_1 = 2
    numero_defecto_columna_2 = 1
    estado_columna1=12

    ventana = Tk()
    ventana.title("Defectos Relacionados")

    texto = Text(ventana, wrap="word")
    texto.pack(fill="both", expand=True)

    scrollbar = Scrollbar(ventana, orient="vertical", command=texto.yview)
    scrollbar.pack(side="right", fill="y")
    texto.config(yscrollcommand=scrollbar.set)

    dict_defectos = {}

    defectos_excel_2 = {cell[0].value.lower(): (hoja_2.cell(row=cell[0].row, column=numero_defecto_columna_2).value, nlp(cell[0].value.lower())) for cell in hoja_2.iter_rows(min_row=2, min_col=descripcion_columna_2, max_col=descripcion_columna_2)}

    for cell in hoja_1.iter_rows(min_row=2, min_col=descripcion_columna_1, max_col=descripcion_columna_1):
        issue = hoja_1.cell(row=cell[0].row, column=numero_defecto_columna_1).value
        estado = hoja_1.cell(row=cell[0].row, column=estado_columna1).value
        if estado == "1. Abierto":
            issue_desc = nlp(cell[0].value.lower())
            defectos_relacionados = []
            for defecto, (numero_defecto, defecto_desc) in defectos_excel_2.items():
                similitud = issue_desc.similarity(defecto_desc)
                porcentaje_similitud = round(similitud * 100, 2)
                if similitud > 0.25:
                    defectos_relacionados.append((numero_defecto, porcentaje_similitud))
            dict_defectos.setdefault(issue, []).extend(defectos_relacionados)

    for issue, defectos in dict_defectos.items():
        texto.insert("end", f"Issue: {issue}\n", "issue_header")
        for defecto in defectos:
            color = 'black'
            if defecto[1] <= 40:
                color = 'red'
                texto.insert("end", f"- {defecto[0]} ({defecto[1]}%) - Poco probable\n", color)
            elif defecto[1] > 70:
                color = 'green'
                texto.insert("end", f"- {defecto[0]} ({defecto[1]}%) - Muy probable\n", color)
            else:
                texto.insert("end", f"- {defecto[0]} ({defecto[1]}%)\n", color)

    texto.tag_configure("issue_header", font=("Arial", 10, "bold"))
    texto.tag_configure("red", foreground="red")
    texto.tag_configure("green", foreground="green")

    def actualizar_tamano(event):
        texto.configure(font=("Arial", max(10, int(ventana.winfo_width() / 80))))

    ventana.bind("<Configure>", actualizar_tamano)

    ventana.mainloop()

root = Tk()
root.title("Defectos")
root.geometry("1000x600")
root.configure(bg="lightblue")

menu = Menu(root)
root.config(menu=menu)

boton_seleccion_excel = Button(root, text="Seleccionar Excel", command=seleccionar_excel, width=60, height=10, bg="light yellow")
boton_seleccion_excel.place(relx=0.5, rely=0.5, anchor="center")

root.mainloop()
